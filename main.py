import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
try:
    wb=load_workbook("time.xlsx")
    sheet=wb.active
except:
    wb=Workbook()
    sheet=wb.active
    sheet.title="time data"
user=input('enter country').lower().replace(" ", "-")
def find_country(sheet,user):
    for row in range(2,sheet.max_row+1):
        if sheet[f"A{row}"]==user:
            return row
    return None
row = find_country(sheet, user)
if not row:
    row = sheet.max_row + 1
    sheet[f"A{row}"] = user
while True:
    sheet['A1'] = 'country'
    sheet['B1'] = 'country time'
    sheet['C1'] = 'local time'
    sheet['D1'] = 'time difference'
    sheet['E1'] = 'country weather'
    for cell in sheet[1]:
        cell.font=Font(size=15,bold=True)
    url_time=requests.get(f"https://www.timeanddate.com/worldclock/{user}")
    #print(url.status_code)
    #url_weather=requests.get(f"https://www.timeanddate.com/weather/{user}")
    soup_time=BeautifulSoup(url_time.text,"html.parser")
    #soup_weather=BeautifulSoup(url_weather.text,"html.parser")
    time_element=soup_time.find("span",id="ct")
    #weather_element=soup_time.find("div",class_="h2")
    country_time=time_element.text
    country_time=country_time.replace("ص", "AM").replace("م", "PM")
    local_time=datetime.now()
    local_time = local_time.strftime("%I:%M:%S %p").lstrip("0")
    country_dt=datetime.strptime(country_time,"%I:%M:%S %p")
    local_dt=datetime.strptime(local_time,"%I:%M:%S %p")
    diff=abs(country_dt-local_dt)
    total_seconds = int(diff.total_seconds())
    hours = total_seconds // 3600
    #minutes = (total_seconds % 3600) // 60
    #seconds = total_seconds % 60
    time_difference=f"{hours} hrs"
    sheet[f"A{row}"]=user
    sheet[f"B{row}"] = country_time
    sheet[f"C{row}"] = local_time
    sheet[f"D{row}"] = time_difference
    wb.save("time.xlsx")
    time.sleep(1)
    print(f"current time in {user} is : {country_time}")
    print(f"current local time is : {local_time}")
    print("---------------------------------")
